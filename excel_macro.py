import openpyxl

def get_user_input(prompt):
    return input(prompt + ": ")

def main():
    file_name = 'sample.xlsx'
    wb = openpyxl.load_workbook(file_name)
    ws = wb['Sheet1']

    max_row = ws.max_row - 1

    # Get input from the user
    a_name = get_user_input("A열 Name 입력")
    a_start_num = int(get_user_input("A열 시작번호"))
    a_input_type = get_user_input("A열 Type 입력")
    
    a_second_input = get_user_input("두번째 줄에 들어갈 것 입력")
    input_chk_type = get_user_input("두번째 줄에 type을 입력할래? (y/n)")
    
    if input_chk_type == 'y':
        data_type = ", type=" + a_input_type.upper()
    else:
        data_type = ""
    
    a_third_input = get_user_input("A열 3행 ex) cover")
    a_third_start_num = int(get_user_input("A열 3행 시작번호"))
    a_chk_third = get_user_input("세번째 줄 입력할거 있는가 (없으면 빈칸 있으면 입력 ex)11,11,)")

    for i in range(1, max_row * 3, 3):
        ws['A' + str(i)] = f'** Name: {a_name}-{a_start_num} Type: {a_input_type}'
        ws['A' + str(i + 1)] = f'*{a_second_input}{data_type}'
        ws['A' + str(i + 2)] = f'{a_third_input}-{a_third_start_num},{a_chk_third}{ws['B' + str(i // 3 + 1)].value}'
        a_start_num += 1

    save_name = get_user_input("저장할 파일명 입력")
    wb.save(f'{save_name}.xlsx')

if __name__ == "__main__":
    main()
